"""
市场规模分析 Excel 模板生成器

生成一个标准化的市场规模分析 Excel 模板，包含：
- Sheet 1: 核心假设 (Assumptions)
- Sheet 2: TAM 计算
- Sheet 3: SAM 计算
- Sheet 4: SOM 计算
- Sheet 5: 敏感性矩阵

使用方法:
    python generate_template.py [输出路径]
"""

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# 样式定义
BLUE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # 输入项
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 标题
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 结果
HEADER_FONT = Font(bold=True, size=11)
INPUT_FONT = Font(color="0000FF", bold=True)  # 蓝色 = 可编辑
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_assumptions_sheet(ws):
    """Sheet 1: 核心假设"""
    ws.title = "核心假设"
    
    # 标题
    ws['A1'] = "市场规模分析 - 核心假设"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # 说明
    ws['A3'] = "💡 说明: 蓝色单元格为可编辑输入项，修改后其他表格自动更新"
    ws['A3'].font = Font(italic=True, color="666666")
    ws.merge_cells('A3:D3')
    
    # 假设表头
    headers = ["假设项", "数值", "单位", "来源/备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = GRAY_FILL
        cell.border = THIN_BORDER
    
    # 示例假设
    assumptions = [
        ("行业总规模", 5000, "亿元", "艾瑞2024报告"),
        ("目标细分占比", 0.08, "%", "8% = HR SaaS 占企业软件"),
        ("地域占比", 0.4, "%", "40% = 华东华北占比"),
        ("目标渗透率", 0.3, "%", "30% 企业已采用"),
        ("年均客单价", 10, "万元", "中大型客户"),
        ("目标市占率", 0.02, "%", "2% = 保守估计"),
        ("预测 CAGR", 0.12, "%", "12% 年复合增长"),
    ]
    
    for row, (name, value, unit, note) in enumerate(assumptions, 6):
        ws.cell(row=row, column=1, value=name).border = THIN_BORDER
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.fill = BLUE_FILL
        value_cell.font = INPUT_FONT
        value_cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=unit).border = THIN_BORDER
        ws.cell(row=row, column=4, value=note).border = THIN_BORDER
    
    # 列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30


def create_tam_sheet(ws):
    """Sheet 2: TAM 计算"""
    ws.title = "TAM计算"
    
    ws['A1'] = "TAM (Total Addressable Market) 总可触达市场"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Top-Down 方法
    ws['A3'] = "方法一: Top-Down"
    ws['A3'].font = HEADER_FONT
    
    ws['A4'] = "行业总规模"
    ws['B4'] = "=核心假设!B6"
    ws['C4'] = "亿元"
    
    ws['A5'] = "TAM (Top-Down)"
    ws.cell(row=5, column=2, value="=B4").fill = GREEN_FILL
    ws['C5'] = "亿元"
    
    # Bottom-Up 方法
    ws['A7'] = "方法二: Bottom-Up"
    ws['A7'].font = HEADER_FONT
    
    ws['A8'] = "目标客户数量"
    ws.cell(row=8, column=2, value=500000).fill = BLUE_FILL
    ws.cell(row=8, column=2).font = INPUT_FONT
    ws['C8'] = "家"
    
    ws['A9'] = "目标渗透率"
    ws['B9'] = "=核心假设!B9"
    
    ws['A10'] = "年均客单价"
    ws['B10'] = "=核心假设!B10"
    ws['C10'] = "万元"
    
    ws['A11'] = "TAM (Bottom-Up)"
    ws.cell(row=11, column=2, value="=B8*B9*B10/10000").fill = GREEN_FILL
    ws['C11'] = "亿元"
    
    # 对比
    ws['A13'] = "TAM 对比"
    ws['A13'].font = HEADER_FONT
    ws['A14'] = "Top-Down"
    ws['B14'] = "=B5"
    ws['A15'] = "Bottom-Up"
    ws['B15'] = "=B11"
    ws['A16'] = "平均值"
    ws.cell(row=16, column=2, value="=(B14+B15)/2").fill = GREEN_FILL
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15


def create_sam_sheet(ws):
    """Sheet 3: SAM 计算"""
    ws.title = "SAM计算"
    
    ws['A1'] = "SAM (Serviceable Addressable Market) 可服务市场"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = "计算过程"
    ws['A3'].font = HEADER_FONT
    
    ws['A4'] = "TAM"
    ws['B4'] = "=TAM计算!B16"
    ws['C4'] = "亿元"
    
    ws['A5'] = "× 目标细分占比"
    ws['B5'] = "=核心假设!B7"
    
    ws['A6'] = "× 地域占比"
    ws['B6'] = "=核心假设!B8"
    
    ws['A8'] = "SAM"
    ws.cell(row=8, column=2, value="=B4*B5*B6").fill = GREEN_FILL
    ws.cell(row=8, column=2).font = Font(bold=True, size=12)
    ws['C8'] = "亿元"
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15


def create_som_sheet(ws):
    """Sheet 4: SOM 计算"""
    ws.title = "SOM计算"
    
    ws['A1'] = "SOM (Serviceable Obtainable Market) 可获取市场"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = "计算过程"
    ws['A3'].font = HEADER_FONT
    
    ws['A4'] = "SAM"
    ws['B4'] = "=SAM计算!B8"
    ws['C4'] = "亿元"
    
    ws['A5'] = "× 目标市占率"
    ws['B5'] = "=核心假设!B11"
    
    ws['A7'] = "SOM"
    ws.cell(row=7, column=2, value="=B4*B5").fill = GREEN_FILL
    ws.cell(row=7, column=2).font = Font(bold=True, size=12)
    ws['C7'] = "亿元"
    
    # 多年预测
    ws['A10'] = "多年预测"
    ws['A10'].font = HEADER_FONT
    
    years = ["2024", "2025", "2026", "2027", "2028"]
    for col, year in enumerate(years, 2):
        ws.cell(row=11, column=col, value=year).font = HEADER_FONT
    
    ws['A12'] = "SOM"
    for col in range(2, 7):
        if col == 2:
            ws.cell(row=12, column=col, value="=B7")
        else:
            prev_col = get_column_letter(col - 1)
            ws.cell(row=12, column=col, value=f"={prev_col}12*(1+核心假设!$B$12)")
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12


def create_sensitivity_sheet(ws):
    """Sheet 5: 敏感性矩阵"""
    ws.title = "敏感性分析"
    
    ws['A1'] = "敏感性分析 - SOM 对关键假设的敏感度"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = "行: 增长率 (CAGR) | 列: 市占率"
    ws['A3'].font = Font(italic=True)
    
    # 矩阵标题
    share_rates = [0.01, 0.02, 0.03, 0.05, 0.08]
    growth_rates = [0.05, 0.10, 0.12, 0.15, 0.20]
    
    ws['A5'] = "CAGR \\ 市占率"
    ws['A5'].fill = GRAY_FILL
    
    for col, rate in enumerate(share_rates, 2):
        cell = ws.cell(row=5, column=col, value=f"{rate:.0%}")
        cell.fill = GRAY_FILL
        cell.font = HEADER_FONT
    
    for row, growth in enumerate(growth_rates, 6):
        cell = ws.cell(row=row, column=1, value=f"{growth:.0%}")
        cell.fill = GRAY_FILL
        cell.font = HEADER_FONT
        
        for col, share in enumerate(share_rates, 2):
            # 简化公式: SAM * share * (1+growth)^3
            ws.cell(row=row, column=col, value=f"=SAM计算!$B$8*{share}*(1+{growth})^3")
    
    ws.column_dimensions['A'].width = 15


def generate_template(output_path: str):
    """生成完整的 Excel 模板"""
    wb = Workbook()
    
    # 创建各 Sheet
    create_assumptions_sheet(wb.active)
    create_tam_sheet(wb.create_sheet())
    create_sam_sheet(wb.create_sheet())
    create_som_sheet(wb.create_sheet())
    create_sensitivity_sheet(wb.create_sheet())
    
    # 保存
    wb.save(output_path)
    print(f"✅ 模板已生成: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        output = sys.argv[1]
    else:
        output = Path(__file__).parent.parent / "templates" / "market_sizing_template.xlsx"
    
    generate_template(str(output))

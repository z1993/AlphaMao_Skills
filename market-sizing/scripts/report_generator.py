"""
Market Sizing Report Generator
==============================

生成专业级市场规模分析报告：
- Markdown 报告 (带 Mermaid 图表)
- HTML 单页报告 (Plotly 交互图表)
- Excel 工作簿 (公式联动)

使用方法:
    from report_generator import ReportGenerator
    
    gen = ReportGenerator()
    gen.generate(
        market_name="中国航空活塞发动机 (200-500HP)",
        tam=25.5,
        sam=11.5,
        som=1.7,
        fermi_result=fermi_result,
        monte_carlo_result=mc_result,
        output_dir="./output"
    )
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Any
from dataclasses import dataclass, asdict

# 尝试导入可视化库
PLOTLY_AVAILABLE = False
try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    pass

# 尝试导入 Excel 库
OPENPYXL_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass


@dataclass
class MarketSizingData:
    """市场规模分析数据"""
    market_name: str                      # 市场名称
    geography: str                        # 地域范围
    base_year: int                        # 基准年份
    forecast_years: int                   # 预测年数
    
    tam: float                            # Total Addressable Market
    sam: float                            # Serviceable Addressable Market
    som: float                            # Serviceable Obtainable Market
    unit: str                             # 单位 (如 "亿元")
    
    cagr: float                           # 复合增长率
    
    # 可选：详细数据
    core_insight: Optional[str] = None    # 核心洞察 (1-2句话)
    market_definition: Optional[dict] = None  # 市场边界 {"产品": {"含": ..., "排除": ...}, ...}
    growth_forecast: Optional[List[dict]] = None  # 年度预测 [{"year": 2025, "tam": X, "sam": Y}]
    fermi_result: Optional[dict] = None   # Fermi 计算结果
    monte_carlo_result: Optional[dict] = None  # Monte Carlo 结果
    assumptions: Optional[List[dict]] = None   # 假设列表 (含 name, value, key, type: 📚/🧮/⚠️)
    competitors: Optional[List[dict]] = None   # 竞争者列表
    data_sources: Optional[List[str]] = None   # 数据来源
    
    # 推导链：每个核心数字怎么来的
    tam_derivation: Optional[dict] = None
    # {"method": "Top-Down", "steps": [{"desc":"..","value":"..","source":"📚.."}], "result": X}
    sam_derivation: Optional[dict] = None
    # {"method": "Bottom-Up Fermi", "ref": "fermi_result", "result": X}
    som_derivation: Optional[dict] = None
    # {"method": "SAM × 目标市占率", "steps": [...], "result": X}
    top_down_result: Optional[dict] = None
    # {"steps": [{"desc":"..","value":X,"source":"📚"},...], "result": X}
    cross_validation: Optional[dict] = None
    # {"bottom_up": X, "top_down": Y, "deviation": "Z%", "conclusion": "..."}
    risks: Optional[List[dict]] = None
    # [{"type":"数据时效性", "detail":"..."}]
    growth_drivers: Optional[List[str]] = None
    # ["健康消费升级", "大瓶装渗透低线城市"]

    def validate(self) -> List[str]:
        """校验数据完整性，返回 warnings 列表。不阻塞生成，但打印告警。"""
        warnings = []
        # 检查核心假设
        if not self.assumptions:
            warnings.append("⚠️ assumptions 为空 → Excel 将无公式链 (全部降级为静态值)")
        else:
            keys = [a.get("key") for a in self.assumptions if a.get("key")]
            missing_fields = []
            for a in self.assumptions:
                if not a.get("key"): missing_fields.append(f"假设 '{a.get('name','')}' 缺少 key")
                if a.get("numeric_value") is None: missing_fields.append(f"假设 '{a.get('name','')}' 缺少 numeric_value")
                if not a.get("type"): missing_fields.append(f"假设 '{a.get('name','')}' 缺少 type (📚/🧮/⚠️)")
            if missing_fields:
                warnings.extend(missing_fields)
            # 检查通用 key
            if "cagr" not in keys:
                warnings.append("⚠️ 缺少 cagr key → Growth Forecast 将使用静态 CAGR 值")
            if "som_share" not in keys:
                warnings.append("⚠️ 缺少 som_share key → SOM 将降级为静态值")
        # 检查 Fermi 相关
        if not self.fermi_result:
            warnings.append("⚠️ fermi_result 为空 → Fermi 静态 fallback 也不可用")
        if not self.monte_carlo_result:
            warnings.append("⚠️ monte_carlo_result 为空 → Monte Carlo Sheet 将显示'未运行'")
        if not self.tam_derivation:
            warnings.append("⚠️ tam_derivation 为空 → MD/HTML 报告缺少 TAM 推导过程")
        return warnings


class ReportGenerator:
    """
    专业报告生成器
    
    支持输出格式：
    - Markdown (.md)
    - HTML (.html) - 带 Plotly 交互图表
    """
    
    def __init__(self):
        self.template_dir = Path(__file__).parent.parent / "templates"
    
    def generate_markdown(
        self,
        data: MarketSizingData,
        output_path: Path
    ) -> str:
        """
        生成 Markdown 报告
        """
        lines = []
        
        # 标题
        lines.append(f"# 市场规模分析报告：{data.market_name}")
        lines.append("")
        lines.append(f"> **市场**: {data.market_name}")
        lines.append(f"> **地域范围**: {data.geography}")
        lines.append(f"> **基准年份**: {data.base_year}")
        lines.append(f"> **分析日期**: {datetime.now().strftime('%Y-%m-%d')}")
        lines.append("")
        lines.append("---")
        lines.append("")
        
        # 执行摘要
        lines.append("## 执行摘要")
        lines.append("")
        lines.append("| 指标 | 数值 | 说明 |")
        lines.append("|------|------|------|")
        lines.append(f"| **TAM** | {data.tam:.1f} {data.unit} | 总可寻址市场 |")
        lines.append(f"| **SAM** | {data.sam:.1f} {data.unit} | 可服务市场 |")
        lines.append(f"| **SOM** | {data.som:.1f} {data.unit} | 可获取市场 |")
        lines.append(f"| **CAGR** | {data.cagr*100:.1f}% | 预计复合增长率 |")
        lines.append("")
        
        # 核心洞察
        if data.core_insight:
            lines.append(f"**核心洞察**: {data.core_insight}")
            lines.append("")
        
        
        # TAM/SAM/SOM 漏斗 (Mermaid)
        lines.append("### 市场层级")
        lines.append("")
        lines.append("```mermaid")
        lines.append("graph TD")
        lines.append(f'    TAM["TAM: {data.tam:.1f} {data.unit}"]')
        lines.append(f'    SAM["SAM: {data.sam:.1f} {data.unit}"]')
        lines.append(f'    SOM["SOM: {data.som:.1f} {data.unit}"]')
        lines.append("    TAM --> SAM")
        lines.append("    SAM --> SOM")
        lines.append("    style TAM fill:#4CAF50,color:#fff")
        lines.append("    style SAM fill:#2196F3,color:#fff")
        lines.append("    style SOM fill:#FF9800,color:#fff")
        lines.append("```")
        lines.append("")
        
        # 市场边界定义
        if data.market_definition:
            lines.append("---")
            lines.append("")
            lines.append("## 市场边界定义")
            lines.append("")
            lines.append("| 维度 | 包含 (In Scope) | 排除 (Out of Scope) |")
            lines.append("|------|-----------------|---------------------|")
            for dim, scope in data.market_definition.items():
                in_scope = scope.get("含", scope.get("in", ""))
                out_scope = scope.get("排除", scope.get("out", "—"))
                lines.append(f"| **{dim}** | {in_scope} | {out_scope} |")
            lines.append("")
        
        # Fermi 分解
        if data.fermi_result:
            lines.append("---")
            lines.append("")
            lines.append("## Fermi 估算分解")
            lines.append("")
            lines.append(f"**方法**: {data.fermi_result.get('method', 'N/A')}")
            lines.append("")
            lines.append(f"**公式**: `{data.fermi_result.get('formula', 'N/A')}`")
            lines.append("")
            
            if data.fermi_result.get("steps"):
                lines.append("### 计算步骤")
                lines.append("")
                lines.append("```mermaid")
                lines.append("graph LR")
                steps = data.fermi_result["steps"]
                for i, (desc, val, basis) in enumerate(steps):
                    node_id = f"S{i}"
                    val_str = self._format_number(val)
                    lines.append(f'    {node_id}["{desc}<br/>{val_str}"]')
                    if i > 0:
                        lines.append(f"    S{i-1} --> {node_id}")
                lines.append("```")
                lines.append("")
        
        # Monte Carlo 结果
        if data.monte_carlo_result:
            lines.append("---")
            lines.append("")
            lines.append("## Monte Carlo 模拟结果")
            lines.append("")
            mc = data.monte_carlo_result
            lines.append(f"**模拟次数**: {mc.get('n_simulations', 10000):,}")
            lines.append("")
            lines.append("### 置信区间")
            lines.append("")
            lines.append("| 分位数 | 数值 |")
            lines.append("|--------|------|")
            lines.append(f"| P5 (悲观) | {self._format_number(mc.get('p5', 0))} {data.unit} |")
            lines.append(f"| P25 | {self._format_number(mc.get('p25', 0))} {data.unit} |")
            lines.append(f"| **P50 (中位数)** | **{self._format_number(mc.get('median', 0))} {data.unit}** |")
            lines.append(f"| P75 | {self._format_number(mc.get('p75', 0))} {data.unit} |")
            lines.append(f"| P95 (乐观) | {self._format_number(mc.get('p95', 0))} {data.unit} |")
            lines.append("")
            
            # 敏感性分析
            if mc.get("sensitivity"):
                lines.append("### 敏感性分析 (Tornado)")
                lines.append("")
                lines.append("| 假设 | 影响幅度 |")
                lines.append("|------|----------|")
                sorted_sens = sorted(mc["sensitivity"].items(), key=lambda x: abs(x[1]), reverse=True)
                for name, impact in sorted_sens:
                    bar = "▓" * min(int(abs(impact) / 5), 10)
                    lines.append(f"| {name} | {bar} {impact:+.1f}% |")
                lines.append("")
        
        # 假设列表
        if data.assumptions:
            lines.append("---")
            lines.append("")
            lines.append("## 核心假设清单")
            lines.append("")
            lines.append("> 📚 = 引用可靠报告 | 🧮 = 由已有数据计算推导 | ⚠️ = 合理假设")
            lines.append("")
            lines.append("| 假设项 | 数值 | 来源/推导 | 类型 |")
            lines.append("|--------|------|-----------|------|")
            for assumption in data.assumptions:
                name = assumption.get("name", "")
                value = assumption.get("value", "")
                source = assumption.get("source", "估算")
                a_type = assumption.get("type", "⚠️")
                lines.append(f"| {name} | {value} | {source} | {a_type} |")
            lines.append("")
        
        # 竞争格局
        if data.competitors:
            lines.append("---")
            lines.append("")
            lines.append("## 竞争格局")
            lines.append("")
            lines.append("| 公司 | 市占率 | 优势 | 数据来源 |")
            lines.append("|------|--------|------|----------|")
            for comp in data.competitors:
                name = comp.get("name", "")
                share = comp.get("market_share", "")
                advantage = comp.get("advantage", "")
                source = comp.get("source", "")
                lines.append(f"| {name} | {share} | {advantage} | {source} |")
            lines.append("")
        
        # 增长预测
        if data.growth_forecast:
            lines.append("---")
            lines.append("")
            lines.append("## 增长预测")
            lines.append("")
            lines.append("| 年份 | TAM | SAM | 增长率 |")
            lines.append("|------|-----|-----|--------|")
            for row in data.growth_forecast:
                year = row.get("year", "")
                tam = row.get("tam", "")
                sam = row.get("sam", "")
                growth = row.get("growth", "—")
                lines.append(f"| {year} | {tam} | {sam} | {growth} |")
            lines.append("")
        
        # 数据来源
        if data.data_sources:
            lines.append("---")
            lines.append("")
            lines.append("## 数据来源")
            lines.append("")
            for i, source in enumerate(data.data_sources, 1):
                lines.append(f"{i}. {source}")
            lines.append("")
        
        # 写入文件
        content = "\n".join(lines)
        output_path.write_text(content, encoding="utf-8")
        
        return content
    
    def generate_html(
        self,
        data: MarketSizingData,
        output_path: Path
    ) -> str:
        """
        生成自包含 HTML 报告 (纯 CSS 可视化, 无外部依赖)
        包含完整 12 个章节, 离线可用
        """
        html_content = self._build_html_report(data)
        output_path.write_text(html_content, encoding="utf-8")
        return html_content

    def _build_html_report(self, data: MarketSizingData) -> str:
        """构建 Notion 风格自包含 HTML 报告 (完整 10 节)"""

        def pct(val, mx):
            return f"{val / mx * 100:.1f}" if mx else "0"

        def _deriv_steps_html(steps):
            """渲染推导链步骤为 HTML"""
            html = ""
            for s in steps:
                src = s.get("source", "")
                tag_cls = "tag-blue" if "📚" in src else "tag-green" if "🧮" in src else "tag-amber"
                tag = src.split(" ")[0] if src else ""
                html += f'''<div class="deriv-step">
                    <div class="deriv-desc">{s.get("desc","")}</div>
                    <div class="deriv-val">{s.get("value","")}</div>
                    <span class="tag {tag_cls}">{tag}</span>
                </div>\n'''
            return html

        # ========== §1 执行摘要 ==========
        # Derivation summary under KPIs
        deriv_summary = ""
        derivs = [
            ("TAM", data.tam_derivation), ("SAM", data.sam_derivation), ("SOM", data.som_derivation)
        ]
        for label, d in derivs:
            if d:
                deriv_summary += f'<div class="deriv-line"><strong>{label}</strong> = {d.get("result","")} {data.unit} <span class="muted">← {d.get("method","")}</span></div>\n'

        # Core insight
        insight_html = f'<div class="callout callout-blue"><span class="callout-icon">💡</span><div>{data.core_insight}</div></div>' if data.core_insight else ""

        # Funnel
        funnel_html = ""
        colors = ["#2ecc71", "#3498db", "#e67e22"]
        labels = [("TAM", data.tam, "总可寻址市场"), ("SAM", data.sam, "可服务市场"), ("SOM", data.som, "可获取市场")]
        for (label, val, desc), color in zip(labels, colors):
            w = pct(val, data.tam)
            funnel_html += f'''<div class="funnel-row">
                <div class="funnel-bar" style="width:{w}%;background:{color}"><span>{label}</span><span>{val:,.1f} {data.unit}</span></div>
                <span class="funnel-desc">{desc}</span></div>\n'''

        # ========== §2 市场边界 ==========
        mkt_def_html = ""
        if data.market_definition:
            rows = ""
            for dim, scope in data.market_definition.items():
                inc = scope.get("含", scope.get("in", ""))
                exc = scope.get("排除", scope.get("out", "—"))
                rows += f"<tr><td><strong>{dim}</strong></td><td>{inc}</td><td class='muted'>{exc}</td></tr>\n"
            mkt_def_html = f'''<section><h2>§2 🎯 市场边界定义</h2>
                <table><thead><tr><th>维度</th><th>包含 (In Scope)</th><th>排除 (Out of Scope)</th></tr></thead>
                <tbody>{rows}</tbody></table></section>'''

        # ========== §3 核心假设 ==========
        assume_html = ""
        if data.assumptions:
            rows = ""
            for a in data.assumptions:
                t = a.get("type", "⚠️")
                cls = "tag-blue" if t == "📚" else "tag-green" if t == "🧮" else "tag-amber"
                used = a.get("used_in", "")
                used_html = f'<span class="muted">→ {used}</span>' if used else ""
                rows += f'''<tr><td>{a.get("name","")}</td><td><strong>{a.get("value","")}</strong></td>
                    <td class="muted">{a.get("source","")}</td><td><span class="tag {cls}">{t}</span></td>
                    <td>{used_html}</td></tr>\n'''
            assume_html = f'''<section><h2>§3 📋 核心假设清单</h2>
                <div class="legend">📚 = 引用可靠报告 &nbsp;&nbsp; 🧮 = 由已有数据计算 &nbsp;&nbsp; ⚠️ = 合理假设</div>
                <table><thead><tr><th>假设项</th><th>数值</th><th>来源 / 推导</th><th>类型</th><th>引用处</th></tr></thead>
                <tbody>{rows}</tbody></table></section>'''

        # ========== §4 双重验证 ==========
        validation_html = ""
        # 4.1 Bottom-Up Fermi Tree
        fermi_tree_html = ""
        if data.fermi_result:
            fr = data.fermi_result
            tree_nodes = ""
            if fr.get("steps"):
                for i, (desc, val, basis) in enumerate(fr["steps"]):
                    is_last = (i == len(fr["steps"]) - 1)
                    node_cls = "tree-result" if is_last else ""
                    connector = "└──" if is_last else "├──"
                    tree_nodes += f'''<div class="tree-node {node_cls}">
                        <span class="tree-connector">{connector}</span>
                        <span class="tree-desc">{desc}</span>
                        <span class="tree-val">{self._format_number(val)}</span>
                    </div>\n'''
            method_info = f'<strong>方法</strong>: {fr.get("method","N/A")} · <code>{fr.get("formula","")}</code>'
            fermi_tree_html = f'''<h3>4.1 Bottom-Up (Fermi 分解)</h3>
                <div class="callout callout-gray"><span class="callout-icon">📐</span><div>{method_info}</div></div>
                <div class="tree-diagram">{tree_nodes}</div>'''

        # 4.2 Top-Down
        topdown_html = ""
        if data.top_down_result:
            td = data.top_down_result
            topdown_html = f'''<h3>4.2 Top-Down</h3>
                <p class="muted">{td.get("method","")}</p>
                <div class="deriv-chain">{_deriv_steps_html(td.get("steps",[]))}</div>
                <div class="callout callout-gray"><span class="callout-icon">🔢</span><div>Top-Down 结果: <strong>{td.get("result","")} {data.unit}</strong></div></div>'''
        elif data.tam_derivation:
            td = data.tam_derivation
            topdown_html = f'''<h3>4.2 Top-Down (TAM 推导)</h3>
                <p class="muted">{td.get("method","")}</p>
                <div class="deriv-chain">{_deriv_steps_html(td.get("steps",[]))}</div>'''

        # 4.3 Cross Validation
        xval_html = ""
        if data.cross_validation:
            cv = data.cross_validation
            xval_html = f'''<h3>4.3 交叉验证</h3>
                <table><thead><tr><th>方法</th><th>结果 ({data.unit})</th><th>偏差</th></tr></thead>
                <tbody>
                    <tr><td>Bottom-Up</td><td><strong>{cv.get("bottom_up","")}</strong></td><td>基准</td></tr>
                    <tr><td>Top-Down</td><td><strong>{cv.get("top_down","")}</strong></td><td>{cv.get("deviation","")}</td></tr>
                </tbody></table>
                <div class="callout callout-blue"><span class="callout-icon">✅</span><div>{cv.get("conclusion","")}</div></div>'''

        if fermi_tree_html or topdown_html or xval_html:
            validation_html = f'<section><h2>§4 🔬 市场测算：双重验证</h2>{fermi_tree_html}{topdown_html}{xval_html}</section>'

        # ========== §5 增长预测 ==========
        forecast_html = ""
        fc = data.growth_forecast
        if not fc:
            sr = data.sam / data.tam if data.tam else 0.8
            fc = []
            for i in range(data.forecast_years + 1):
                yr = data.base_year + i
                t = round(data.tam * (1 + data.cagr) ** i, 1)
                s = round(t * sr, 1)
                g = f"{data.cagr*100:.0f}%" if i > 0 else "—"
                fc.append({"year": yr, "tam": t, "sam": s, "growth": g})
        if fc:
            rows = ""
            mx = max(float(str(r.get("tam", 0)).replace(",", "")) for r in fc) if fc else 1
            bars = ""
            for r in fc:
                tv = float(str(r.get("tam", 0)).replace(",", ""))
                sv = float(str(r.get("sam", 0)).replace(",", "")) if r.get("sam") else 0
                rows += f'<tr><td>{r.get("year","")}</td><td>{r.get("tam","")}</td><td>{r.get("sam","")}</td><td>{r.get("growth","—")}</td></tr>\n'
                bars += f'''<div class="bar-col"><div class="bar-pair">
                    <div class="bar bar-t" style="height:{tv/mx*100:.0f}px" title="TAM {r.get('tam','')}"></div>
                    <div class="bar bar-s" style="height:{sv/mx*100:.0f}px" title="SAM {r.get('sam','')}"></div>
                    </div><div class="bar-yr">{r.get("year","")}</div></div>\n'''

            # Growth drivers
            drivers_html = ""
            if data.growth_drivers:
                items = "".join(f"<li>{d}</li>" for d in data.growth_drivers)
                drivers_html = f'<h3>增长驱动</h3><ul class="driver-list">{items}</ul>'

            forecast_html = f'''<section><h2>§5 📈 增长预测 (CAGR {data.cagr*100:.0f}%)</h2>
                <div class="chart-legend"><span class="dot dot-t"></span> TAM <span class="dot dot-s"></span> SAM</div>
                <div class="bar-chart">{bars}</div>
                <table><thead><tr><th>年份</th><th>TAM ({data.unit})</th><th>SAM ({data.unit})</th><th>增长率</th></tr></thead>
                <tbody>{rows}</tbody></table>{drivers_html}</section>'''

        # ========== §6 Monte Carlo ==========
        mc_html = ""
        if data.monte_carlo_result:
            mc = data.monte_carlo_result
            pills = ""
            for label, key, hl in [("P5 悲观","p5",""), ("P25","p25",""), ("P50 中位数","median","hl"), ("P75","p75",""), ("P95 乐观","p95","")]:
                pills += f'<div class="pill {hl}"><div class="pill-label">{label}</div><div class="pill-val">{self._format_number(mc.get(key,0))} {data.unit}</div></div>\n'
            mc_html = f'''<section><h2>§6 🎲 Monte Carlo 模拟</h2>
                <p>模拟次数: <strong>{mc.get("n_simulations",10000):,}</strong></p>
                <div class="pill-grid">{pills}</div></section>'''

        # ========== §7 敏感性分析 ==========
        sens_html = ""
        if data.monte_carlo_result and data.monte_carlo_result.get("sensitivity"):
            mc = data.monte_carlo_result
            sorted_s = sorted(mc["sensitivity"].items(), key=lambda x: abs(x[1]), reverse=True)
            mx_impact = max(abs(v) for _, v in sorted_s) if sorted_s else 1
            tornado = ""
            sens_rows = ""
            for i, (name, impact) in enumerate(sorted_s):
                w = abs(impact) / mx_impact * 100
                tornado += f'''<div class="tornado-row">
                    <div class="tornado-name">{name}</div>
                    <div class="tornado-track"><div class="tornado-fill" style="width:{w:.0f}%"></div></div>
                    <div class="tornado-val">{impact:+.1f}%</div></div>\n'''
                # 解读
                interp = "最关键变量" if i == 0 else "次关键" if i == 1 else "影响有限"
                sens_rows += f'<tr><td>{name}</td><td><strong>{impact:+.1f}%</strong></td><td class="muted">{interp}</td></tr>\n'

            top_var = sorted_s[0][0] if sorted_s else ""
            second_var = sorted_s[1][0] if len(sorted_s) > 1 else ""
            ratio = round(sorted_s[0][1] / sorted_s[1][1], 1) if len(sorted_s) > 1 and sorted_s[1][1] else ""
            insight_text = f"{top_var}的影响是{second_var}的{ratio}倍" if ratio else ""

            sens_html = f'''<section><h2>§7 📊 敏感性分析</h2>
                {tornado}
                <table><thead><tr><th>假设</th><th>影响幅度</th><th>解读</th></tr></thead>
                <tbody>{sens_rows}</tbody></table>
                {"<div class='callout callout-amber'><span class='callout-icon'>⚡</span><div>" + insight_text + "</div></div>" if insight_text else ""}
                </section>'''

        # ========== §8 竞争格局 ==========
        comp_html = ""
        if data.competitors:
            rows = ""
            total_share = 0
            for c in data.competitors:
                rows += f'''<tr><td><strong>{c.get("name","")}</strong></td><td>{c.get("market_share","")}</td>
                    <td>{c.get("advantage","")}</td><td class="muted">{c.get("source","")}</td></tr>\n'''
                try:
                    total_share += float(c.get("market_share","0").replace("%",""))
                except:
                    pass
            cr_note = ""
            if total_share > 0:
                concentration = "高度集中" if total_share >= 60 else "中度集中" if total_share >= 40 else "碎片化"
                cr_note = f'<p><strong>CR{len(data.competitors)}</strong>: {total_share:.0f}% → {concentration}</p>'
            comp_html = f'''<section><h2>§8 🏆 竞争格局</h2>
                <table><thead><tr><th>品牌 / 企业</th><th>市占率</th><th>核心优势</th><th>数据来源</th></tr></thead>
                <tbody>{rows}</tbody></table>{cr_note}</section>'''

        # ========== §9 风险与局限性 ==========
        risk_html = ""
        if data.risks:
            items = ""
            for r in data.risks:
                items += f'<div class="risk-item"><strong>{r.get("type","")}</strong>: {r.get("detail","")}</div>\n'
            risk_html = f'<section><h2>§9 ⚠️ 风险与局限性</h2>{items}</section>'

        # ========== §10 数据来源 ==========
        src_html = ""
        if data.data_sources:
            items = "".join(f"<li>{s}</li>" for s in data.data_sources)
            src_html = f'<section><h2>§10 📚 数据来源汇总</h2><ol>{items}</ol></section>'

        # ========== ASSEMBLE HTML ==========
        html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{data.market_name} — 市场规模分析</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Serif+SC:wght@400;700&family=Inter:wght@400;500;600&display=swap');
:root{{--bg:#fbfbfa;--container:#fff;--text:#37352f;--muted:#9b9a97;--border:#e9e9e7;
--blue:#2383e2;--blue-bg:#e8f0fe;--green:#0f7b6c;--green-bg:#dff4f0;--amber:#d9730d;--amber-bg:#fbecdd;
--gray-bg:#f1f1ef;--radius:6px}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Inter',-apple-system,'PingFang SC','Microsoft YaHei',sans-serif;
background:var(--bg);color:var(--text);font-size:15px;line-height:1.75;-webkit-font-smoothing:antialiased}}
.page{{max-width:860px;margin:0 auto;background:var(--container);min-height:100vh;padding:0 clamp(24px,5vw,96px)}}

.cover{{padding:72px 0 40px;border-bottom:1px solid var(--border)}}
.cover h1{{font-family:'Noto Serif SC',Georgia,serif;font-size:2.4rem;font-weight:700;line-height:1.3;margin-bottom:8px}}
.cover .sub{{color:var(--muted);font-size:.95rem}}

section{{padding:32px 0;border-bottom:1px solid var(--border)}}
section:last-of-type{{border-bottom:none}}
h2{{font-family:'Noto Serif SC',Georgia,serif;font-size:1.35rem;font-weight:700;margin-bottom:16px;color:var(--text)}}
h3{{font-size:1rem;font-weight:600;margin:20px 0 10px;color:var(--text)}}
p{{margin-bottom:12px;color:#555}}
code{{background:var(--gray-bg);padding:2px 6px;border-radius:4px;font-size:.85em;color:var(--amber)}}

.kpi-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px}}
.kpi{{background:var(--gray-bg);border-radius:var(--radius);padding:16px;text-align:center}}
.kpi-label{{font-size:.7rem;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);font-weight:500}}
.kpi-value{{font-size:1.7rem;font-weight:700;margin:4px 0}}
.kpi:nth-child(1) .kpi-value{{color:var(--green)}} .kpi:nth-child(2) .kpi-value{{color:var(--blue)}}
.kpi:nth-child(3) .kpi-value{{color:var(--amber)}} .kpi:nth-child(4) .kpi-value{{color:#6940a5}}

/* Derivation summary */
.deriv-summary{{background:var(--gray-bg);border-radius:var(--radius);padding:12px 16px;margin:12px 0}}
.deriv-line{{font-size:.9rem;margin-bottom:4px}}
.deriv-line .muted{{font-size:.8rem}}

/* Derivation chain */
.deriv-chain{{margin:8px 0}}
.deriv-step{{display:flex;align-items:center;gap:8px;padding:6px 0;border-bottom:1px solid var(--border)}}
.deriv-step:last-child{{border-bottom:none}}
.deriv-desc{{flex:1;font-size:.9rem}}
.deriv-val{{font-weight:600;font-size:.9rem;min-width:80px;text-align:right}}

.callout{{display:flex;align-items:flex-start;gap:10px;padding:14px 16px;border-radius:var(--radius);margin:12px 0;font-size:.95rem}}
.callout-icon{{font-size:1.1rem;flex-shrink:0;margin-top:2px}}
.callout-blue{{background:var(--blue-bg);color:#1a5276}}
.callout-gray{{background:var(--gray-bg);color:#555}}
.callout-amber{{background:var(--amber-bg);color:#7e4a1e}}

.funnel-row{{display:flex;align-items:center;gap:12px;margin-bottom:8px}}
.funnel-bar{{height:36px;border-radius:var(--radius);display:flex;align-items:center;justify-content:space-between;
padding:0 14px;min-width:100px;color:#fff;font-size:.85rem;font-weight:500}}
.funnel-desc{{color:var(--muted);font-size:.8rem}}

table{{width:100%;border-collapse:collapse;font-size:.9rem;margin:8px 0}}
thead th{{background:var(--gray-bg);color:var(--muted);font-weight:500;font-size:.75rem;
text-transform:uppercase;letter-spacing:.04em;padding:8px 12px;text-align:left;border-bottom:1px solid var(--border)}}
td{{padding:10px 12px;border-bottom:1px solid var(--border);vertical-align:top}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:#fafaf8}}
.muted{{color:var(--muted);font-size:.85rem}}

.tag{{display:inline-block;padding:1px 8px;border-radius:4px;font-size:.8rem;font-weight:500}}
.tag-blue{{background:var(--blue-bg);color:var(--blue)}}
.tag-green{{background:var(--green-bg);color:var(--green)}}
.tag-amber{{background:var(--amber-bg);color:var(--amber)}}
.legend{{color:var(--muted);font-size:.8rem;margin-bottom:10px}}

/* Fermi Tree */
.tree-diagram{{font-family:'Consolas','Courier New',monospace;background:var(--gray-bg);
border-radius:var(--radius);padding:16px 20px;margin:12px 0;font-size:.9rem;line-height:2}}
.tree-node{{display:flex;align-items:center;gap:8px;padding:2px 0}}
.tree-connector{{color:var(--muted);font-weight:600;min-width:30px}}
.tree-desc{{flex:1;color:#555}}
.tree-val{{font-weight:700;min-width:100px;text-align:right}}
.tree-result{{border-top:2px solid var(--border);margin-top:4px;padding-top:6px}}
.tree-result .tree-val{{color:var(--blue);font-size:1.05rem}}

/* Pills */
.pill-grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin:12px 0}}
.pill{{background:var(--gray-bg);border-radius:var(--radius);padding:10px;text-align:center}}
.pill.hl{{background:var(--blue-bg);border:1.5px solid rgba(35,131,226,0.3)}}
.pill-label{{font-size:.7rem;color:var(--muted);margin-bottom:2px}}
.pill-val{{font-size:.85rem;font-weight:600}}
.pill.hl .pill-val{{color:var(--blue);font-size:.95rem}}

/* Tornado */
.tornado-row{{display:flex;align-items:center;gap:10px;margin-bottom:6px}}
.tornado-name{{width:130px;flex-shrink:0;font-size:.85rem;color:var(--muted);text-align:right}}
.tornado-track{{flex:1;height:20px;background:var(--gray-bg);border-radius:4px;overflow:hidden}}
.tornado-fill{{height:100%;background:linear-gradient(90deg,var(--blue),#5dade2);border-radius:4px}}
.tornado-val{{width:55px;font-size:.85rem;font-weight:600;color:var(--blue)}}

/* Bar Chart */
.bar-chart{{display:flex;align-items:flex-end;gap:16px;justify-content:center;padding:24px 0 8px}}
.bar-col{{display:flex;flex-direction:column;align-items:center}}
.bar-pair{{display:flex;gap:3px;align-items:flex-end}}
.bar{{width:24px;border-radius:4px 4px 0 0}}
.bar-t{{background:var(--green)}} .bar-s{{background:var(--blue)}}
.bar-yr{{font-size:.75rem;color:var(--muted);margin-top:4px}}
.chart-legend{{display:flex;gap:16px;font-size:.8rem;color:var(--muted);margin-bottom:4px}}
.dot{{width:10px;height:10px;border-radius:2px;display:inline-block;margin-right:2px}}
.dot-t{{background:var(--green)}} .dot-s{{background:var(--blue)}}

/* Driver / Risk */
.driver-list{{padding-left:1.5rem;color:#555;font-size:.9rem;margin-top:8px}}
.driver-list li{{margin-bottom:6px}}
.risk-item{{background:var(--gray-bg);border-radius:var(--radius);padding:10px 14px;margin-bottom:8px;font-size:.9rem;color:#555}}

ol{{padding-left:1.5rem;color:#555;font-size:.9rem}} ol li{{margin-bottom:4px}}
ul{{padding-left:1.5rem}}

.footer{{text-align:center;padding:32px 0;color:var(--muted);font-size:.8rem}}

@media(max-width:640px){{
    .kpi-row,.pill-grid{{grid-template-columns:repeat(2,1fr)}}
    .tornado-name{{width:80px;font-size:.75rem}}
    .bar-chart{{gap:8px}} .bar{{width:16px}}
}}
</style>
</head>
<body>
<div class="page">
    <div class="cover">
        <h1>{data.market_name}</h1>
        <div class="sub">{data.geography} · {data.base_year} · 市场规模分析报告</div>
    </div>

    <section>
        <h2>§1 📊 执行摘要</h2>
        <div class="kpi-row">
            <div class="kpi"><div class="kpi-label">TAM</div><div class="kpi-value">{data.tam:,.1f}</div><div class="kpi-label">{data.unit}</div></div>
            <div class="kpi"><div class="kpi-label">SAM</div><div class="kpi-value">{data.sam:,.1f}</div><div class="kpi-label">{data.unit}</div></div>
            <div class="kpi"><div class="kpi-label">SOM</div><div class="kpi-value">{data.som:,.1f}</div><div class="kpi-label">{data.unit}</div></div>
            <div class="kpi"><div class="kpi-label">CAGR</div><div class="kpi-value">{data.cagr*100:.0f}%</div><div class="kpi-label">复合增长率</div></div>
        </div>
        {"<div class='deriv-summary'>" + deriv_summary + "</div>" if deriv_summary else ""}
        {insight_html}
        <h3>TAM / SAM / SOM</h3>
        {funnel_html}
    </section>

    {mkt_def_html}
    {assume_html}
    {validation_html}
    {forecast_html}
    {mc_html}
    {sens_html}
    {comp_html}
    {risk_html}
    {src_html}

    <div class="footer">报告生成于 {datetime.now().strftime("%Y-%m-%d %H:%M")} · Market Sizing Skill</div>
</div>
</body>
</html>'''
        return html




    
    def generate(
        self,
        data: MarketSizingData,
        output_dir: Path,
        formats: List[str] = ["md", "html"]
    ) -> Dict[str, Path]:
        """
        生成所有格式的报告
        
        Args:
            data: 市场规模数据
            output_dir: 输出目录
            formats: 输出格式列表 ["md", "html", "xlsx"]
            
        Returns:
            生成的文件路径字典
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        safe_name = data.market_name.replace(" ", "_").replace("/", "-")[:50]
        timestamp = datetime.now().strftime("%Y%m%d")
        base_name = f"market_sizing_{safe_name}_{timestamp}"
        
        results = {}
        
        # 校验数据完整性
        warnings = data.validate()
        if warnings:
            print("\n🔍 数据完整性检查:")
            for w in warnings:
                print(f"  {w}")
            print()
        
        if "md" in formats:
            md_path = output_dir / f"{base_name}.md"
            self.generate_markdown(data, md_path)
            results["md"] = md_path
        
        if "html" in formats:
            html_path = output_dir / f"{base_name}.html"
            try:
                self.generate_html(data, html_path)
                results["html"] = html_path
            except ImportError as e:
                print(f"⚠️ 跳过 HTML 生成: {e}")
        
        if "xlsx" in formats:
            xlsx_path = output_dir / f"{base_name}.xlsx"
            try:
                self.generate_excel(data, xlsx_path)
                results["xlsx"] = xlsx_path
            except ImportError as e:
                print(f"⚠️ 跳过 Excel 生成: {e}")
        
        return results
    
    def generate_excel(
        self,
        data: MarketSizingData,
        output_path: Path
    ) -> str:
        """
        生成 Excel 分析模型 (5 个 Sheet)
        特点: 假设驱动 (Assumption-Driven)，全公式联动
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        wb = Workbook()
        
        # ── 样式定义 (遵循 xlsx skill 标准) ──
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        title_font = Font(size=14, bold=True, color="1F4E79")
        # xlsx skill 颜色编码
        input_font = Font(color="0000FF")       # 蓝字 = 硬编码输入/假设
        formula_font = Font(color="000000")      # 黑字 = 公式计算
        xref_font = Font(color="008000")         # 绿字 = 跨表引用
        assumption_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄底 = 关键假设
        calc_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # 公式结果区
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        def _write_header(ws, row, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        def _write_row(ws, row, values, fills=None, formats=None, fonts=None):
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                if fills and col in fills:
                    cell.fill = fills[col]
                if formats and col in formats:
                    cell.number_format = formats[col]
                if fonts and col in fonts:
                    cell.font = fonts[col]

        def _set_col_widths(ws, widths):
            for col, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = w
        
        # =========================================================
        # Sheet 1: 核心假设 (INPUTS) — 蓝字+黄底
        # =========================================================
        ws1 = wb.active
        ws1.title = "核心假设"
        
        ws1.merge_cells('A1:E1')
        ws1['A1'] = f"{data.market_name} — 市场规模模型"
        ws1['A1'].font = title_font
        ws1['A2'] = "修改 B 列蓝色数值，后续 Sheet 结果自动更新。蓝字=输入 | 黑字=公式 | 绿字=跨表引用"
        ws1['A2'].font = Font(italic=True, color="666666")

        _write_header(ws1, 4, ["假设项 (Key)", "数值 (Input)", "单位/说明", "来源 (Source)", "类型"])
        
        # 建立 Key -> Cell 映射 (e.g., "pene_rate": "'核心假设'!B6")
        key_map = {} 
        
        if data.assumptions:
            for i, a in enumerate(data.assumptions):
                row_idx = 5 + i
                val = a.get("numeric_value", a.get("value"))
                
                # 数字格式：用 key 名判断百分比，而非纯数值范围
                a_key = a.get("key", "")
                is_pct = any(hint in a_key for hint in ("rate", "pct", "ratio", "share", "adopt", "cagr"))
                if is_pct and isinstance(val, float):
                    fmt = '0.0%'
                elif isinstance(val, (int, float)) and abs(val) >= 1000:
                    fmt = '#,##0'
                else:
                    fmt = '#,##0.0'
                
                # 数据类型决定样式
                dtype = a.get("type", "⚠️")
                # ⚠️ 假设 → 蓝字+黄底，📚 引用 → 蓝字，🧮 计算 → 黑字
                if dtype == "⚠️":
                    b_fill = assumption_fill
                    b_font = input_font
                elif dtype == "🧮":
                    b_fill = calc_fill
                    b_font = formula_font
                else:
                    b_fill = None
                    b_font = input_font
                
                fills_dict = {}
                if b_fill:
                    fills_dict[2] = b_fill
                
                _write_row(ws1, row_idx, [
                    a.get("name", ""),
                    val,
                    a.get("value", ""),
                    a.get("source", ""),
                    dtype,
                ], fills=fills_dict, formats={2: fmt}, fonts={2: b_font})
                
                if a.get("key"):
                    key_map[a["key"]] = f"'核心假设'!B{row_idx}"
        
        _set_col_widths(ws1, [30, 15, 20, 35, 8])

        # =========================================================
        # Sheet 2: Fermi 计算 (CALC ENGINE) — 通用分段模型
        # =========================================================
        ws2 = wb.create_sheet("Fermi_Calc")
        ws2['A1'] = "Bottom-Up Fermi 计算过程"
        ws2['A1'].font = title_font
        ws2['A2'] = "绿字=引用核心假设 | 黑字=公式 | 所有中间步骤可追溯"
        ws2['A2'].font = Font(italic=True, color="666666")
        
        _write_header(ws2, 3, ["步骤", "说明", f"数值/公式 ({data.unit})", "来源类型"])
        
        fermi_final_cell = None
        
        # ── 自动检测分段模式 ──
        # Pattern 1: institution_based segments (e.g. ka_count/ka_adopt/ka_price)
        # Pattern 2: population_based chain (base_pop/core_pop_pct/pene_rate/freq/price)
        
        # 检测 institution_based 分段
        seg_prefixes = []
        seg_keys_found = {}
        for k in key_map:
            # 匹配 count/vol/num/institutions 等后缀
            for suffix in ("_count", "_vol", "_num", "_institutions", "_number"):
                if k.endswith(suffix):
                    prefix = k[:-len(suffix)]
                    # 查找同前缀的 adopt/rate/penetration 和 price/cost/arpu
                    adopt_key = None
                    price_key = None
                    for ak in key_map:
                        if not ak.startswith(prefix):
                            continue
                        ak_suffix = ak[len(prefix):]
                        if any(h in ak_suffix for h in ("adopt", "rate", "pene", "penetration")):
                            adopt_key = ak
                        if any(h in ak_suffix for h in ("price", "cost", "arpu", "spend", "revenue")):
                            price_key = ak
                    if adopt_key and price_key:
                        seg_keys_found[prefix] = {
                            "count": k,
                            "adopt": adopt_key,
                            "price": price_key,
                        }
        
        is_institution_based = len(seg_keys_found) > 0
        is_population_based = all(k in key_map for k in ["base_pop", "pene_rate", "price"])
        # Pattern 3: substitution_based (existing_market + substitution_rate + price_premium)
        is_substitution = all(k in key_map for k in ["existing_market", "substitution_rate"])
        # Pattern 4: value_chain_based (end_market + value_share)
        is_value_chain = all(k in key_map for k in ["end_market", "value_share"])
        # Pattern 5: value_based (target_count + prob_freq + prob_cost + wtp_ratio)
        is_value_based = all(k in key_map for k in ["target_count", "prob_cost", "wtp_ratio"])
        
        curr_row = 4
        
        if is_institution_based:
            # ── Institution-Based 分段模型 ──
            segment_result_cells = []
            
            for seg_prefix, seg_keys in seg_keys_found.items():
                seg_label = seg_prefix.upper()
                # 从假设中找到对应的中文名
                for a in (data.assumptions or []):
                    if a.get("key") == seg_keys["count"]:
                        seg_label = a.get("used_in", seg_prefix.upper()).split()[-1] if a.get("used_in") else seg_prefix.upper()
                        break
                
                # 段标题
                ws2.cell(row=curr_row, column=1, value=f"── {seg_label} 段 ──")
                ws2.cell(row=curr_row, column=1).font = Font(bold=True, size=11)
                curr_row += 1
                
                # Row: 机构数
                count_ref = key_map[seg_keys["count"]]
                ws2.cell(row=curr_row, column=1, value=f"{seg_label}: 机构数")
                ws2.cell(row=curr_row, column=2, value="引用假设")
                ws2.cell(row=curr_row, column=3, value=f"={count_ref}")
                ws2.cell(row=curr_row, column=3).font = xref_font
                ws2.cell(row=curr_row, column=4, value="🧮 引用")
                for c in range(1, 5):
                    ws2.cell(row=curr_row, column=c).border = thin_border
                count_cell = f"C{curr_row}"
                curr_row += 1
                
                # Row: 有效机构 = 机构数 × 渗透率
                adopt_ref = key_map[seg_keys["adopt"]]
                ws2.cell(row=curr_row, column=1, value=f"{seg_label}: 有效机构")
                ws2.cell(row=curr_row, column=2, value="× 渗透率/采用率")
                ws2.cell(row=curr_row, column=3, value=f"={count_cell}*{adopt_ref}")
                ws2.cell(row=curr_row, column=3).font = xref_font
                ws2.cell(row=curr_row, column=4, value="🧮 计算")
                for c in range(1, 5):
                    ws2.cell(row=curr_row, column=c).border = thin_border
                adopt_cell = f"C{curr_row}"
                curr_row += 1
                
                # Row: 市场规模 = 有效机构 × 客单价
                price_ref = key_map[seg_keys["price"]]
                ws2.cell(row=curr_row, column=1, value=f"{seg_label}: 市场规模")
                ws2.cell(row=curr_row, column=2, value="× 客单价")
                ws2.cell(row=curr_row, column=3, value=f"={adopt_cell}*{price_ref}")
                ws2.cell(row=curr_row, column=3).font = xref_font
                ws2.cell(row=curr_row, column=3).number_format = '#,##0.00'
                ws2.cell(row=curr_row, column=4, value=data.unit)
                for c in range(1, 5):
                    ws2.cell(row=curr_row, column=c).border = thin_border
                segment_result_cells.append(f"C{curr_row}")
                curr_row += 1
                
                # 空行
                curr_row += 1
            
            # TAM 汇总行
            ws2.cell(row=curr_row, column=1, value="★ TAM (Bottom-Up 汇总)")
            ws2.cell(row=curr_row, column=1).font = Font(bold=True, size=12, color="1F4E79")
            ws2.cell(row=curr_row, column=2, value="= 各段之和")
            sum_formula = "+".join(segment_result_cells)
            ws2.cell(row=curr_row, column=3, value=f"={sum_formula}")
            ws2.cell(row=curr_row, column=3).font = Font(bold=True, size=12)
            ws2.cell(row=curr_row, column=3).number_format = '#,##0.00'
            ws2.cell(row=curr_row, column=3).fill = calc_fill
            ws2.cell(row=curr_row, column=4, value=data.unit)
            for c in range(1, 5):
                ws2.cell(row=curr_row, column=c).border = thin_border
            fermi_final_cell = f"C{curr_row}"
        
        elif is_population_based:
            # ── Population-Based 链式模型 ──
            # numeric_value 单位约定: base_pop 以亿为单位, price 以元为单位
            # 最终换算到 data.unit (通常是"亿元")
            # unit_scale: 人口(亿) × 占比 × 渗透率 × 频次 × 单价(元) = 亿元
            # 无需额外乘除，只要 base_pop 是亿人、price 是元
            prev_cell = None
            
            if "base_pop" in key_map:
                # base_pop numeric_value 是亿人, 需要转为人
                f_val = f"={key_map['base_pop']}*100000000"
                _write_row(ws2, 4, ["L1: 基础人口", "引用假设 (亿→人)", f_val, "人"],
                           fonts={3: xref_font})
                prev_cell = "C4"
            
            if "core_pop_pct" in key_map and prev_cell:
                f_val = f"={prev_cell}*{key_map['core_pop_pct']}"
                _write_row(ws2, 5, ["L2: 核心人群", "× 人群占比", f_val, "人"],
                           fonts={3: xref_font})
                prev_cell = "C5"
                
            if "pene_rate" in key_map and prev_cell:
                f_val = f"={prev_cell}*{key_map['pene_rate']}"
                _write_row(ws2, 6, ["L3: 消费用户数", "× 渗透率", f_val, "人"],
                           fonts={3: xref_font})
                prev_cell = "C6"

            if "freq" in key_map and prev_cell:
                f_val = f"={prev_cell}*{key_map['freq']}"
                _write_row(ws2, 7, ["L4: 总消费量", "× 频次", f_val, "次"],
                           fonts={3: xref_font})
                prev_cell = "C7"
                
            if "price" in key_map and prev_cell:
                # price 是元, 结果转亿元: / 1e8
                f_val = f"={prev_cell}*{key_map['price']}/100000000"
                _write_row(ws2, 8, ["L5: 市场规模", f"× 单价 → {data.unit}", f_val, data.unit],
                           fonts={3: xref_font})
                fermi_final_cell = "C8"
                ws2["C8"].fill = calc_fill
                ws2["C8"].number_format = '#,##0.00'
                curr_row = 9
        
        elif is_substitution:
            # ── 替代法: existing_market × substitution_rate × price_premium ──
            ws2.cell(row=4, column=1, value="L1: 现有市场规模")
            ws2.cell(row=4, column=2, value="引用假设")
            ws2.cell(row=4, column=3, value=f"={key_map['existing_market']}")
            ws2.cell(row=4, column=3).font = xref_font
            for c in range(1, 5): ws2.cell(row=4, column=c).border = thin_border
            
            ws2.cell(row=5, column=1, value="L2: 替代率")
            ws2.cell(row=5, column=2, value="× 替代率")
            ws2.cell(row=5, column=3, value=f"=C4*{key_map['substitution_rate']}")
            ws2.cell(row=5, column=3).font = xref_font
            for c in range(1, 5): ws2.cell(row=5, column=c).border = thin_border
            
            if "price_premium" in key_map:
                ws2.cell(row=6, column=1, value="L3: 价格调整")
                ws2.cell(row=6, column=2, value="× 价格系数")
                ws2.cell(row=6, column=3, value=f"=C5*{key_map['price_premium']}")
                ws2.cell(row=6, column=3).font = xref_font
                ws2.cell(row=6, column=3).fill = calc_fill
                ws2.cell(row=6, column=3).number_format = '#,##0.00'
                for c in range(1, 5): ws2.cell(row=6, column=c).border = thin_border
                fermi_final_cell = "C6"
            else:
                ws2["C5"].fill = calc_fill
                ws2["C5"].number_format = '#,##0.00'
                fermi_final_cell = "C5"
            curr_row = 8
        
        elif is_value_chain:
            # ── 价值链法: end_market × value_share ──
            ws2.cell(row=4, column=1, value="L1: 终端市场规模")
            ws2.cell(row=4, column=2, value="引用假设")
            ws2.cell(row=4, column=3, value=f"={key_map['end_market']}")
            ws2.cell(row=4, column=3).font = xref_font
            for c in range(1, 5): ws2.cell(row=4, column=c).border = thin_border
            
            ws2.cell(row=5, column=1, value="L2: 环节价值占比")
            ws2.cell(row=5, column=2, value="× 价值占比")
            ws2.cell(row=5, column=3, value=f"=C4*{key_map['value_share']}")
            ws2.cell(row=5, column=3).font = xref_font
            ws2.cell(row=5, column=3).fill = calc_fill
            ws2.cell(row=5, column=3).number_format = '#,##0.00'
            for c in range(1, 5): ws2.cell(row=5, column=c).border = thin_border
            fermi_final_cell = "C5"
            curr_row = 7
        
        elif is_value_based:
            # ── 价值基础法: target_count × prob_freq × prob_cost × wtp_ratio ──
            ws2.cell(row=4, column=1, value="L1: 目标客户数")
            ws2.cell(row=4, column=2, value="引用假设")
            ws2.cell(row=4, column=3, value=f"={key_map['target_count']}")
            ws2.cell(row=4, column=3).font = xref_font
            for c in range(1, 5): ws2.cell(row=4, column=c).border = thin_border
            prev = "C4"
            
            if "prob_freq" in key_map:
                ws2.cell(row=5, column=1, value="L2: 问题频率")
                ws2.cell(row=5, column=2, value="× 问题频率")
                ws2.cell(row=5, column=3, value=f"={prev}*{key_map['prob_freq']}")
                ws2.cell(row=5, column=3).font = xref_font
                for c in range(1, 5): ws2.cell(row=5, column=c).border = thin_border
                prev = "C5"
            
            ws2.cell(row=6, column=1, value="L3: × 问题成本")
            ws2.cell(row=6, column=2, value="× 单次成本")
            ws2.cell(row=6, column=3, value=f"={prev}*{key_map['prob_cost']}")
            ws2.cell(row=6, column=3).font = xref_font
            for c in range(1, 5): ws2.cell(row=6, column=c).border = thin_border
            
            ws2.cell(row=7, column=1, value="L4: × 愿付比例")
            ws2.cell(row=7, column=2, value="× WTP")
            ws2.cell(row=7, column=3, value=f"=C6*{key_map['wtp_ratio']}")
            ws2.cell(row=7, column=3).font = xref_font
            ws2.cell(row=7, column=3).fill = calc_fill
            ws2.cell(row=7, column=3).number_format = '#,##0.00'
            for c in range(1, 5): ws2.cell(row=7, column=c).border = thin_border
            fermi_final_cell = "C7"
            curr_row = 9
        
        # 兜底：如果所有模式都不匹配，打印静态 steps
        if not fermi_final_cell and data.fermi_result and data.fermi_result.get("steps"):
            steps = data.fermi_result["steps"]
            for i, step in enumerate(steps):
                if isinstance(step, (list, tuple)):
                    desc, val, basis = step[0], step[1], step[2] if len(step) > 2 else ""
                elif isinstance(step, dict):
                    desc = step.get("desc", "")
                    val = step.get("value", "")
                    basis = step.get("source", "")
                else:
                    continue
                _write_row(ws2, 4+i, [desc, basis, val, "静态值"],
                           fonts={3: Font(color="FF0000")})  # 红字标记无公式

        _set_col_widths(ws2, [28, 25, 25, 15])
        
        # =========================================================
        # Sheet 3: TAM / SAM / SOM (OUTPUTS) — 全公式
        # =========================================================
        ws3 = wb.create_sheet("Market_Size_Output")
        ws3['A1'] = "TAM / SAM / SOM 结果看板"
        ws3['A1'].font = title_font
        ws3['A2'] = "所有数值由公式计算，修改假设表自动更新"
        ws3['A2'].font = Font(italic=True, color="666666")
        
        _write_header(ws3, 3, ["指标", f"数值 ({data.unit})", "单位", "计算逻辑"])
        
        # TAM: 引用 Fermi 汇总
        if fermi_final_cell:
            ws3["A4"] = "TAM"
            ws3["B4"] = f"=Fermi_Calc!{fermi_final_cell}"
            ws3["B4"].font = xref_font  # 绿字 = 跨表引用
            ws3["B4"].fill = calc_fill
            ws3["B4"].number_format = '#,##0.0'
            ws3["C4"] = data.unit
            ws3["D4"] = f"引用 Fermi_Calc!{fermi_final_cell}"
        else:
            _write_row(ws3, 4, ["TAM", data.tam, data.unit, "静态值 (Fermi未构建)"],
                       fonts={2: Font(color="FF0000")})
        
        # SAM: TAM * sam_ratio 或引用 Fermi
        if "sam_ratio" in key_map:
            ws3["A5"] = "SAM"
            ws3["B5"] = f"=B4*{key_map['sam_ratio']}"
            ws3["B5"].font = xref_font
            ws3["B5"].fill = calc_fill
            ws3["B5"].number_format = '#,##0.0'
            ws3["C5"] = data.unit
            ws3["D5"] = "TAM x 可服务比例"
        elif fermi_final_cell:
            ws3["A5"] = "SAM"
            ws3["B5"] = f"=Fermi_Calc!{fermi_final_cell}"
            ws3["B5"].font = xref_font
            ws3["B5"].fill = calc_fill
            ws3["C5"] = data.unit
            ws3["D5"] = "引用 Fermi 计算结果"
        else:
            _write_row(ws3, 5, ["SAM", data.sam, data.unit, "静态值"])
            
        # SOM: SAM * som_share
        if "som_share" in key_map:
            ws3["A6"] = "SOM"
            ws3["B6"] = f"=B5*{key_map['som_share']}"
            ws3["B6"].font = xref_font
            ws3["B6"].fill = calc_fill
            ws3["B6"].number_format = '#,##0.0'
            ws3["C6"] = data.unit
            ws3["D6"] = "SAM x 目标市占率"
        else:
            _write_row(ws3, 6, ["SOM", data.som, data.unit, "静态值"])
        
        # 添加边框
        for r in range(4, 7):
            for c in range(1, 5):
                ws3.cell(row=r, column=c).border = thin_border
            
        _set_col_widths(ws3, [15, 20, 10, 30])
        
        # =========================================================
        # Sheet 4: 增长预测 (Timeline) — 公式驱动
        # =========================================================
        ws4 = wb.create_sheet("Growth_Forecast")
        ws4['A1'] = "5年增长预测"
        ws4['A1'].font = title_font
        ws4['A2'] = "基准年引用 Market_Size_Output，后续年份用 CAGR 公式推导"
        ws4['A2'].font = Font(italic=True, color="666666")
        
        _write_header(ws4, 3, ["年份", f"TAM ({data.unit})", f"SAM ({data.unit})", "计算逻辑"])
        
        # Year 0 (Base) — 跨表引用，绿字
        base_y = data.base_year
        ws4.cell(row=4, column=1, value=base_y)
        ws4.cell(row=4, column=2, value="=Market_Size_Output!B4")
        ws4.cell(row=4, column=2).font = xref_font
        ws4.cell(row=4, column=2).number_format = '#,##0.0'
        ws4.cell(row=4, column=3, value="=Market_Size_Output!B5")
        ws4.cell(row=4, column=3).font = xref_font
        ws4.cell(row=4, column=3).number_format = '#,##0.0'
        ws4.cell(row=4, column=4, value="基准年 (引用 Market_Size_Output)")
        for c in range(1, 5):
            ws4.cell(row=4, column=c).border = thin_border
        
        # Future Years — 公式
        cagr_ref = key_map.get("cagr", str(data.cagr))
        
        for i in range(1, data.forecast_years + 1):
            r = 4 + i
            ws4.cell(row=r, column=1, value=base_y + i)
            ws4.cell(row=r, column=2, value=f"=B{r-1}*(1+{cagr_ref})")
            ws4.cell(row=r, column=2).font = formula_font
            ws4.cell(row=r, column=2).number_format = '#,##0.0'
            ws4.cell(row=r, column=3, value=f"=C{r-1}*(1+{cagr_ref})")
            ws4.cell(row=r, column=3).font = formula_font
            ws4.cell(row=r, column=3).number_format = '#,##0.0'
            ws4.cell(row=r, column=4, value="=B{prev}*(1+CAGR)".format(prev=r-1))
            for c in range(1, 5):
                ws4.cell(row=r, column=c).border = thin_border

        _set_col_widths(ws4, [10, 20, 20, 35])
        
        # =========================================================
        # Sheet 5: Monte Carlo — 输入假设 + 模拟结果
        # =========================================================
        ws5 = wb.create_sheet("Monte_Carlo")
        
        ws5.merge_cells('A1:D1')
        ws5['A1'] = "Monte Carlo 模拟"
        ws5['A1'].font = title_font
        
        if data.monte_carlo_result:
            mc = data.monte_carlo_result
            
            # ── 输入假设区 ──
            ws5['A3'] = "输入假设 (PERT 三角分布)"
            ws5['A3'].font = Font(bold=True, size=12)
            
            _write_header(ws5, 4, ["假设变量", "最小值", "最可能值", "最大值"])
            mc_assumptions = mc.get('assumptions', {})
            mc_row = 5
            if mc_assumptions:
                for name, vals in mc_assumptions.items():
                    if isinstance(vals, dict):
                        _write_row(ws5, mc_row, [
                            name,
                            vals.get('min', ''),
                            vals.get('most_likely', ''),
                            vals.get('max', ''),
                        ], fonts={2: input_font, 3: input_font, 4: input_font})
                        mc_row += 1
            
            # ── 模拟结果区 ──
            mc_row += 1
            ws5.cell(row=mc_row, column=1, value=f"模拟结果 (n={mc.get('n_simulations', 10000):,})")
            ws5.cell(row=mc_row, column=1).font = Font(bold=True, size=12)
            mc_row += 1
            
            _write_header(ws5, mc_row, ["分位数", f"数值 ({data.unit})", "含义"])
            mc_row += 1
            percentiles = [
                ("P5 (悲观)", mc.get('p5', 0), "5% 概率低于此值"),
                ("P25", mc.get('p25', 0), "下四分位"),
                ("P50 (中位数)", mc.get('median', 0), "最可能结果"),
                ("P75", mc.get('p75', 0), "上四分位"),
                ("P95 (乐观)", mc.get('p95', 0), "5% 概率高于此值"),
            ]
            for i, (label, val, meaning) in enumerate(percentiles):
                _write_row(ws5, mc_row + i, [label, val, meaning])
                ws5.cell(row=mc_row + i, column=2).number_format = '#,##0.00'
            
            # P50 加粗
            ws5.cell(row=mc_row + 2, column=2).font = Font(bold=True, size=12, color="1F4E79")
            
            # ── 敏感性分析区 ──
            if mc.get("sensitivity"):
                sens_row = mc_row + len(percentiles) + 1
                ws5.cell(row=sens_row, column=1, value="敏感性分析 (Tornado)")
                ws5.cell(row=sens_row, column=1).font = Font(bold=True, size=12)
                sens_row += 1
                
                _write_header(ws5, sens_row, ["假设", "影响幅度 (%)", "解读"])
                sens_row += 1
                sorted_sens = sorted(mc["sensitivity"].items(), key=lambda x: abs(x[1]), reverse=True)
                for i, (name, impact) in enumerate(sorted_sens):
                    rank = "最关键变量" if i == 0 else "次关键" if i == 1 else "影响有限"
                    _write_row(ws5, sens_row + i, [name, impact, rank])
                    ws5.cell(row=sens_row + i, column=2).number_format = '+#,##0.0%;-#,##0.0%' if abs(impact) < 1 else '+#,##0.0'
        else:
            ws5['A3'] = "未运行 Monte Carlo 模拟"
            ws5['A3'].font = Font(italic=True, color="999999")
        
        _set_col_widths(ws5, [22, 18, 18, 18])
        
        # ========== Sheet 6: 竞争格局 ==========
        ws6 = wb.create_sheet("竞争格局")
        
        ws6.merge_cells('A1:D1')
        ws6['A1'] = "竞争格局分析"
        ws6['A1'].font = title_font
        
        if data.competitors:
            _write_header(ws6, 3, ["品牌/企业", "市占率", "核心优势", "数据来源"])
            for i, comp in enumerate(data.competitors):
                _write_row(ws6, 4 + i, [
                    comp.get("name", ""),
                    comp.get("market_share", ""),
                    comp.get("advantage", ""),
                    comp.get("source", ""),
                ])
            
            # CR3
            cr_row = 4 + len(data.competitors) + 1
            shares = []
            for c in data.competitors:
                s = c.get("market_share", "")
                if isinstance(s, (int, float)):
                    shares.append(s)
                elif isinstance(s, str):
                    try:
                        shares.append(float(s.replace("%", "")))
                    except ValueError:
                        pass
            if len(shares) >= 3:
                cr3 = sum(sorted(shares, reverse=True)[:3])
                ws6.cell(row=cr_row, column=1, value="CR3").font = Font(bold=True)
                ws6.cell(row=cr_row, column=2, value=f"{cr3:.1f}%")
                concentration = "高度集中" if cr3 > 60 else "中度集中" if cr3 > 30 else "碎片化"
                ws6.cell(row=cr_row, column=3, value=concentration)
        else:
            ws6['A3'] = "未提供竞争数据"
            ws6['A3'].font = Font(italic=True, color="999999")
        
        _set_col_widths(ws6, [22, 12, 28, 22])
        
        # 保存
        wb.save(output_path)
        return str(output_path)
    
    @staticmethod
    def _format_number(n: float) -> str:
        """格式化数字"""
        if n >= 1e12:
            return f"{n/1e12:.2f}万亿"
        elif n >= 1e8:
            return f"{n/1e8:.2f}亿"
        elif n >= 1e4:
            return f"{n/1e4:.2f}万"
        else:
            return f"{n:.2f}"


if __name__ == "__main__":
    # 测试报告生成
    gen = ReportGenerator()
    
    data = MarketSizingData(
        market_name="中国航空活塞发动机 (200-500HP)",
        geography="中国大陆",
        base_year=2024,
        forecast_years=5,
        tam=25.5,
        sam=11.5,
        som=1.7,
        unit="亿元",
        cagr=0.08,
        assumptions=[
            {"name": "市场总规模", "value": "25.5亿元", "source": "IndexBox", "confidence": "中"},
            {"name": "200-500HP占比", "value": "45%", "source": "机型分析", "confidence": "中"},
        ],
        data_sources=[
            "IndexBox: China Aircraft Engine Market Report 2024",
            "Asian Sky Group: China GA Fleet Report",
        ]
    )
    
    output_dir = Path(__file__).parent.parent / "assets"
    results = gen.generate(data, output_dir, formats=["md"])
    
    print("生成的报告:")
    for fmt, path in results.items():
        print(f"  {fmt}: {path}")

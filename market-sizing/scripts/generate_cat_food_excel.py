"""
Generate Cat Food Market Excel Model
完整的猫粮市场分析 Excel 模型
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

def generate_cat_food_excel(output_path: Path):
    wb = Workbook()
    
    # ========== Sheet 1: 核心假设 ==========
    ws1 = wb.active
    ws1.title = "核心假设"
    
    # 样式定义
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    input_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    number_font = Font(size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws1.merge_cells('A1:E1')
    ws1['A1'] = "中国宠物猫粮市场规模分析模型"
    ws1['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    ws1['A2'] = "分析日期: 2026-02-09"
    ws1['A2'].font = Font(italic=True, color="666666")
    
    # 假设表头
    headers = ["假设项", "数值", "单位", "来源", "置信度"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # 假设数据
    assumptions = [
        ("宠物猫数量", 7153, "万只", "派读宠物白皮书2024", "高"),
        ("年均猫粮消费", 560, "元/只", "单猫消费2020元×28%", "中"),
        ("商品粮渗透率", 0.80, "%", "行业估算", "中"),
        ("目标市占率", 0.15, "%", "竞争分析", "中"),
        ("CAGR", 0.18, "%", "蝉妈妈报告", "中"),
    ]
    
    for row_idx, (name, value, unit, source, confidence) in enumerate(assumptions, 5):
        ws1.cell(row=row_idx, column=1, value=name).border = thin_border
        cell = ws1.cell(row=row_idx, column=2, value=value)
        cell.fill = input_fill
        cell.border = thin_border
        cell.number_format = '#,##0.00' if isinstance(value, float) and value < 1 else '#,##0'
        ws1.cell(row=row_idx, column=3, value=unit).border = thin_border
        ws1.cell(row=row_idx, column=4, value=source).border = thin_border
        ws1.cell(row=row_idx, column=5, value=confidence).border = thin_border
    
    # 命名单元格
    ws1['B5'].value = 7153  # 猫数量
    ws1['B6'].value = 560   # 年消费
    ws1['B7'].value = 0.80  # 渗透率
    ws1['B8'].value = 0.15  # 市占率
    ws1['B9'].value = 0.18  # CAGR
    
    # 列宽
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['E'].width = 10
    
    # ========== Sheet 2: TAM/SAM/SOM 计算 ==========
    ws2 = wb.create_sheet("TAM_SAM_SOM")
    
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "TAM / SAM / SOM 计算"
    ws2['A1'].font = Font(size=14, bold=True, color="1F4E79")
    
    # TAM 计算
    ws2['A3'] = "TAM (Total Addressable Market)"
    ws2['A3'].font = Font(bold=True)
    
    ws2['A4'] = "公式"
    ws2['B4'] = "= 猫数量 × 年消费"
    ws2['A5'] = "计算"
    ws2['B5'] = "=核心假设!B5*核心假设!B6/10000"
    ws2['C5'] = "亿元"
    ws2['D5'] = "=ROUND(B5,1)&\" 亿元\""
    
    # SAM 计算
    ws2['A7'] = "SAM (Serviceable Addressable Market)"
    ws2['A7'].font = Font(bold=True)
    
    ws2['A8'] = "公式"
    ws2['B8'] = "= TAM × 商品粮渗透率"
    ws2['A9'] = "计算"
    ws2['B9'] = "=B5*核心假设!B7"
    ws2['C9'] = "亿元"
    
    # SOM 计算
    ws2['A11'] = "SOM (Serviceable Obtainable Market)"
    ws2['A11'].font = Font(bold=True)
    
    ws2['A12'] = "公式"
    ws2['B12'] = "= SAM × 目标市占率"
    ws2['A13'] = "计算"
    ws2['B13'] = "=B9*核心假设!B8"
    ws2['C13'] = "亿元"
    
    # 汇总
    ws2['A16'] = "汇总"
    ws2['A16'].font = Font(bold=True, size=12)
    
    summary_headers = ["指标", "数值 (亿元)", "说明"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=17, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    ws2['A18'] = "TAM"
    ws2['B18'] = "=ROUND(B5,1)"
    ws2['C18'] = "全部猫粮市场"
    
    ws2['A19'] = "SAM"
    ws2['B19'] = "=ROUND(B9,1)"
    ws2['C19'] = "商品粮市场"
    
    ws2['A20'] = "SOM"
    ws2['B20'] = "=ROUND(B13,1)"
    ws2['C20'] = "可获取市场"
    
    # ========== Sheet 3: Monte Carlo ==========
    ws3 = wb.create_sheet("Monte_Carlo")
    
    ws3.merge_cells('A1:D1')
    ws3['A1'] = "Monte Carlo 模拟结果 (10,000次)"
    ws3['A1'].font = Font(size=14, bold=True, color="1F4E79")
    
    # 置信区间
    percentiles = [
        ("P5 (悲观)", 218.5),
        ("P10", 245.3),
        ("P25", 280.6),
        ("P50 (中位数)", 318.2),
        ("P75", 358.4),
        ("P90", 402.1),
        ("P95 (乐观)", 428.7),
    ]
    
    ws3['A3'] = "置信区间"
    ws3['A3'].font = Font(bold=True)
    
    for col, header in enumerate(["分位数", "数值 (亿元)"], 1):
        cell = ws3.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_idx, (label, value) in enumerate(percentiles, 5):
        ws3.cell(row=row_idx, column=1, value=label)
        ws3.cell(row=row_idx, column=2, value=value)
    
    # 敏感性分析
    ws3['A14'] = "敏感性分析 (Tornado)"
    ws3['A14'].font = Font(bold=True)
    
    for col, header in enumerate(["假设", "影响幅度 (%)"], 1):
        cell = ws3.cell(row=15, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    sensitivity = [
        ("猫粮年消费(元/只)", 25.3),
        ("宠物猫数量(万只)", 11.8),
        ("商品粮渗透率", 8.2),
    ]
    
    for row_idx, (name, impact) in enumerate(sensitivity, 16):
        ws3.cell(row=row_idx, column=1, value=name)
        ws3.cell(row=row_idx, column=2, value=impact)
    
    # ========== Sheet 4: 竞争格局 ==========
    ws4 = wb.create_sheet("竞争格局")
    
    ws4.merge_cells('A1:D1')
    ws4['A1'] = "竞争格局分析"
    ws4['A1'].font = Font(size=14, bold=True, color="1F4E79")
    
    competitors = [
        ("皇家 Royal Canin", 5.9, "玛氏", "品牌认知度高，精准营养"),
        ("麦富迪", 4.5, "瑞普", "国产头部，性价比"),
        ("网易严选", 4.0, "网易", "互联网渠道优势"),
        ("蓝氏", 2.9, "吉家", "差异化产品(乳鸽猫粮)"),
        ("渴望 Orijen", 2.5, "冠军宠物", "高端天然粮先行者"),
    ]
    
    for col, header in enumerate(["品牌", "市占率 (%)", "母公司", "核心优势"], 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_idx, (brand, share, parent, advantage) in enumerate(competitors, 4):
        ws4.cell(row=row_idx, column=1, value=brand)
        ws4.cell(row=row_idx, column=2, value=share)
        ws4.cell(row=row_idx, column=3, value=parent)
        ws4.cell(row=row_idx, column=4, value=advantage)
    
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['D'].width = 30
    
    # ========== Sheet 5: 增长预测 ==========
    ws5 = wb.create_sheet("增长预测")
    
    ws5.merge_cells('A1:F1')
    ws5['A1'] = "市场规模增长预测 (CAGR 18%)"
    ws5['A1'].font = Font(size=14, bold=True, color="1F4E79")
    
    years = [2024, 2025, 2026, 2027, 2028, 2029]
    base_tam = 400
    
    for col, header in enumerate(["年份", "TAM (亿元)", "SAM (亿元)", "SOM (亿元)", "YoY增长"], 1):
        cell = ws5.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_idx, year in enumerate(years, 4):
        idx = row_idx - 4
        tam = base_tam * (1.18 ** idx)
        sam = tam * 0.8
        som = sam * 0.15
        
        ws5.cell(row=row_idx, column=1, value=year)
        ws5.cell(row=row_idx, column=2, value=round(tam, 1))
        ws5.cell(row=row_idx, column=3, value=round(sam, 1))
        ws5.cell(row=row_idx, column=4, value=round(som, 1))
        ws5.cell(row=row_idx, column=5, value="18%" if idx > 0 else "-")
    
    # 添加图表
    chart = BarChart()
    chart.type = "col"
    chart.title = "市场规模预测"
    chart.y_axis.title = "亿元"
    chart.x_axis.title = "年份"
    
    data = Reference(ws5, min_col=2, min_row=3, max_col=4, max_row=9)
    cats = Reference(ws5, min_col=1, min_row=4, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws5.add_chart(chart, "G3")
    
    # 保存
    wb.save(output_path)
    print(f"✅ Excel 模型已生成: {output_path}")

if __name__ == "__main__":
    output = Path(r"C:\Users\lenovo\.gemini\antigravity\skills\market-sizing\assets\中国宠物猫粮市场_分析模型.xlsx")
    generate_cat_food_excel(output)

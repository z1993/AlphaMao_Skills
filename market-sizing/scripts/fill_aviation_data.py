"""
Generate Aviation Piston Engine Market Sizing Excel
"""
import sys
from pathlib import Path

# Add skill path to sys.path to access local modules if needed, or just straightforwardly use openpyxl
# Here we will modify the template directly.

try:
    from openpyxl import load_workbook
except ImportError:
    print("错误: 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

def fill_aviation_market_data(template_path, output_path):
    wb = load_workbook(template_path)
    
    # ---------------------------------------------------------
    # Sheet 1: 核心假设 (Assumptions)
    # ---------------------------------------------------------
    ws = wb["核心假设"]
    
    # Update assumptions based on research
    # Data points:
    # 1. Total China Piston Engine Market: $354M (2024) -> ~25.5亿 RMB
    # 2. 200-500HP Segment share: Global is dominant, China estimate 40-50% (Lycoming O-540 is 230-300HP)
    # 3. Target Audience: Flight schools, private owners, sterile UAVs
    # 4. Growth: 1.8亿美元 revenue for piston aircraft by 2026? 
    #    Let's use CAGR from report: 354M -> 398M by 2035 is slow (1.1%). 
    #    But low-altitude economy is booming (26% growth in Wuhu). Let's take a balanced CAGR of 8%.
    
    updates = [
        (6, "中国活塞发动机总规模", 25.5, "亿元", "IndexBox 2024 ($354M)"),
        (7, "200-500HP 细分占比", 0.45, "%", "主力机型(SR20/R44)占比推算"),
        (8, "国产化/区域占比", 1.0, "%", "分析中国整体市场"),
        (9, "潜在替代渗透率", 0.20, "%", "重油/混合动力替代潜力"),
        (10, "发动机均价", 40, "万元", "Lycoming/Continental均价估算"),
        (11, "目标市占率", 0.15, "%", "新进入者保守目标"),
        (12, "预测 CAGR", 0.08, "%", "低空经济政策拉动"), 
        # Add more context rows if needed, but template has fixed rows based on generate_template.py
        # We will just overwrite existing cells for simplicity.
    ]
    
    for row_idx, name, val, unit, note in updates:
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=val)
        ws.cell(row=row_idx, column=3, value=unit)
        ws.cell(row=row_idx, column=4, value=note)

    # ---------------------------------------------------------
    # Sheet 2: TAM 计算
    # ---------------------------------------------------------
    # Top-Down is linked to assumptions, so it auto-updates.
    # Bottom-Up needs manual input in this script
    
    ws_tam = wb["TAM计算"]
    
    # Bottom-up logic:
    # Target Aircraft Fleet: ~3751 piston engines in 2020. 
    # Let's estimate 2024 fleet: ~5000 active piston aircraft.
    # New deliveries + replacements. 
    # Let's say addressable includes new builds + overhauls.
    # 5000 fleet * 10% replacement/overhaul = 500 engines/year
    # + New aircraft sales: 300-500/year?
    # Let's estimate annual demand (units) for Bottom-Up = 1000 engines (aggressive) or 600 (conservative).
    # Using 800 units.
    
    ws_tam.cell(row=8, column=2, value=800) # 目标客户数量 (改为年需求量 units)
    ws_tam.cell(row=8, column=3, value="台/年")
    
    ws_tam.cell(row=9, column=2, value=1.0) # 渗透率 (Addressable is 100% of this demand)
    
    # Unit price linked to assumptions (40万)
    
    # ---------------------------------------------------------
    # Sheet 3, 4, 5
    # ---------------------------------------------------------
    # Auto-calculated based on formulas
    
    wb.save(output_path)
    print(f"✅ Excel 分析文件已生成: {output_path}")

if __name__ == "__main__":
    template = Path(r"C:\Users\lenovo\.gemini\antigravity\skills\market-sizing\templates\market_sizing_template.xlsx")
    output = Path(r"C:\Users\lenovo\.gemini\antigravity\skills\market-sizing\assets\China_Aviation_Piston_Engine_200-500HP.xlsx")
    output.parent.mkdir(exist_ok=True)
    
    fill_aviation_market_data(template, output)
